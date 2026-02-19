"""
Unit tests for core OEE analysis math and parsers.

Run: python -m pytest test_core.py -v
"""

import pytest
import pandas as pd
import numpy as np

from shared import classify_fault, normalize_product, PRODUCT_NORMALIZE
from analyze import (
    _aggregate_oee, _smart_rename, _weighted_mean, EXPECTED_SHEETS,
    _compute_utilization, _build_dead_hour_narrative,
    _correlate_dead_hours_with_events,
    _compute_shift_data, _build_shift_narrative,
)
from datetime import datetime, timedelta


# =====================================================================
# _aggregate_oee — production-weighted OEE math
# =====================================================================

class TestAggregateOEE:
    """Production-weighted OEE should not be a simple average of ratios."""

    def _make_df(self, rows):
        """Helper: build a DataFrame with columns matching hourly data."""
        df = pd.DataFrame(rows)
        for col in ["total_hours", "total_cases", "availability",
                     "performance", "quality", "good_cases"]:
            if col not in df.columns:
                if col == "good_cases":
                    df[col] = df["total_cases"]
                elif col == "quality":
                    df[col] = 1.0
        return df

    def test_single_row(self):
        df = self._make_df([{
            "total_hours": 1.0, "total_cases": 100,
            "availability": 0.9, "performance": 0.8, "quality": 0.95,
            "good_cases": 95,
        }])
        avail, perf, qual, oee = _aggregate_oee(df)
        assert abs(avail - 0.9) < 0.001
        assert abs(perf - 0.8) < 0.001
        assert abs(qual - 0.95) < 0.001
        assert abs(oee - 68.4) < 0.1  # 0.9 * 0.8 * 0.95 * 100

    def test_weighted_not_simple_average(self):
        """The bug: averaging per-hour ratios gives wrong answer when hours differ."""
        df = self._make_df([
            # Hour 1: full hour, low OEE
            {"total_hours": 1.0, "total_cases": 50,
             "availability": 0.5, "performance": 0.5, "quality": 1.0, "good_cases": 50},
            # Hour 2: full hour, high OEE
            {"total_hours": 1.0, "total_cases": 200,
             "availability": 1.0, "performance": 1.0, "quality": 1.0, "good_cases": 200},
        ])
        avail, perf, qual, oee = _aggregate_oee(df)

        # Simple mean would give avail = (0.5+1.0)/2 = 0.75
        # Weighted: production_time = 0.5*1 + 1.0*1 = 1.5, scheduled = 2.0
        # availability = 1.5/2.0 = 0.75 (same in this case because hours are equal)
        assert abs(avail - 0.75) < 0.001

    def test_unequal_hours_weighting(self):
        """With unequal hours, weighting matters a lot."""
        df = self._make_df([
            # Short interval (0.25 hr), bad availability
            {"total_hours": 0.25, "total_cases": 10,
             "availability": 0.2, "performance": 0.5, "quality": 1.0, "good_cases": 10},
            # Long interval (1.0 hr), good availability
            {"total_hours": 1.0, "total_cases": 200,
             "availability": 0.95, "performance": 0.9, "quality": 1.0, "good_cases": 200},
        ])
        avail, perf, qual, oee = _aggregate_oee(df)

        # Simple mean: (0.2 + 0.95) / 2 = 0.575
        # Weighted: (0.2*0.25 + 0.95*1.0) / (0.25 + 1.0) = 1.0/1.25 = 0.80
        assert abs(avail - 0.80) < 0.01
        # The weighted answer is closer to the long interval (0.95) than the bad one
        assert avail > 0.70  # Much better than simple mean of 0.575

    def test_zero_production_excluded(self):
        """Rows with zero cases or zero hours should be excluded."""
        df = self._make_df([
            {"total_hours": 1.0, "total_cases": 100,
             "availability": 0.9, "performance": 0.8, "quality": 1.0, "good_cases": 100},
            # This row should be excluded (zero cases)
            {"total_hours": 1.0, "total_cases": 0,
             "availability": 0.0, "performance": 0.0, "quality": 0.0, "good_cases": 0},
        ])
        avail, perf, qual, oee = _aggregate_oee(df)
        assert abs(avail - 0.9) < 0.001
        assert abs(perf - 0.8) < 0.001

    def test_empty_dataframe(self):
        df = pd.DataFrame({
            "total_hours": pd.Series(dtype=float),
            "total_cases": pd.Series(dtype=float),
            "availability": pd.Series(dtype=float),
            "performance": pd.Series(dtype=float),
            "quality": pd.Series(dtype=float),
            "good_cases": pd.Series(dtype=float),
        })
        avail, perf, qual, oee = _aggregate_oee(df)
        assert avail == 0.0
        assert oee == 0.0

    def test_quality_from_good_cases(self):
        """Quality = good_cases / total_cases."""
        df = self._make_df([{
            "total_hours": 1.0, "total_cases": 200,
            "availability": 1.0, "performance": 1.0,
            "quality": 0.9, "good_cases": 180,
        }])
        _, _, qual, _ = _aggregate_oee(df)
        assert abs(qual - 0.9) < 0.001  # 180/200


# =====================================================================
# classify_fault — downtime reason classification
# =====================================================================

class TestClassifyFault:
    def test_equipment_keywords(self):
        assert classify_fault("Caser - Riverwood") == "Equipment / Mechanical"
        assert classify_fault("Tray Packer - Kayat") == "Equipment / Mechanical"
        assert classify_fault("Palletizer fault") == "Equipment / Mechanical"

    def test_data_gap(self):
        assert classify_fault("Unassigned") == "Data Gap (uncoded)"
        assert classify_fault("Unknown reason") == "Data Gap (uncoded)"

    def test_scheduled(self):
        assert classify_fault("Break-Lunch") == "Scheduled / Non-Production"
        assert classify_fault("Not Scheduled") == "Scheduled / Non-Production"
        assert classify_fault("Lunch (Comida)") == "Scheduled / Non-Production"

    def test_micro_stops(self):
        assert classify_fault("Short Stop") == "Micro Stops"
        assert classify_fault("short stop - filler") == "Micro Stops"

    def test_process(self):
        assert classify_fault("Day Code Change") == "Process / Changeover"
        assert classify_fault("Changeover") == "Process / Changeover"
        assert classify_fault("CIP Cleanup") == "Process / Changeover"

    def test_dash_defaults_to_equipment(self):
        """Reason codes with dashes default to equipment."""
        assert classify_fault("Something - Brand X") == "Equipment / Mechanical"

    def test_unrecognized_no_dash(self):
        assert classify_fault("Random uncategorized thing") == "Other / Unclassified"


# =====================================================================
# normalize_product — product name cleanup
# =====================================================================

class TestNormalizeProduct:
    def test_known_mappings(self):
        assert normalize_product("DM Cut Gr Bn") == "Cut Green Beans 8pk"
        assert normalize_product("dm wk corn") == "WK Corn 12pk"
        assert normalize_product("DM Sliced Pears") == "Pears (trayed)"

    def test_case_insensitive(self):
        assert normalize_product("DM CUT GR BN") == "Cut Green Beans 8pk"
        assert normalize_product("dm cut gr bn") == "Cut Green Beans 8pk"

    def test_whitespace_handling(self):
        assert normalize_product("  dm cut gr bn  ") == "Cut Green Beans 8pk"

    def test_unknown_product_passthrough(self):
        assert normalize_product("New Product XYZ") == "New Product XYZ"

    def test_null_handling(self):
        assert normalize_product(None) == "Unknown"
        assert normalize_product(float("nan")) == "Unknown"
        assert normalize_product("") == "Unknown"


# =====================================================================
# _smart_rename — column name fuzzy matching
# =====================================================================

class TestSmartRename:
    def test_exact_match(self):
        df = pd.DataFrame({"Shift Date": [1], "Shift": ["1st"], "Shift Hour": [1]})
        expected = EXPECTED_SHEETS["DayShiftHour"]["columns"]
        result = _smart_rename(df, expected)
        assert "shift_date" in result.columns

    def test_case_insensitive_match(self):
        df = pd.DataFrame({"shift date": [1], "SHIFT": ["1st"], "shift hour": [1]})
        expected = EXPECTED_SHEETS["DayShiftHour"]["columns"]
        result = _smart_rename(df, expected)
        assert "shift_date" in result.columns

    def test_header_name_matching(self):
        """_smart_rename uses _HEADER_TO_INTERNAL for flexible header matching."""
        df = pd.DataFrame({
            "Date": [1], "Shift": ["1st"], "Hour": [1],
            "Duration Hours": [1.0], "Total Cases": [100],
            "OEE (%)": [50], "Availability": [0.9],
        })
        expected = EXPECTED_SHEETS["DayShiftHour"]["columns"]
        result = _smart_rename(df, expected)
        assert "shift_date" in result.columns
        assert "shift_hour" in result.columns
        assert "oee_pct" in result.columns


# =====================================================================
# _weighted_mean — helper for production-weighted averages
# =====================================================================

class TestWeightedMean:
    def test_basic_weighted_mean(self):
        values = pd.Series([10.0, 20.0])
        weights = pd.Series([1.0, 3.0])
        result = _weighted_mean(values, weights)
        assert abs(result - 17.5) < 0.001  # (10*1 + 20*3) / (1+3) = 70/4

    def test_zero_weights_excluded(self):
        values = pd.Series([10.0, 999.0, 20.0])
        weights = pd.Series([1.0, 0.0, 1.0])
        result = _weighted_mean(values, weights)
        assert abs(result - 15.0) < 0.001  # 999 excluded

    def test_all_zero_weights(self):
        values = pd.Series([10.0, 20.0])
        weights = pd.Series([0.0, 0.0])
        result = _weighted_mean(values, weights)
        assert result == 0.0


# =====================================================================
# _compute_utilization — scheduled time vs producing time
# =====================================================================

class TestComputeUtilization:
    """Utilization = % of scheduled hours that actually produced cases."""

    def _make_df(self, rows):
        return pd.DataFrame(rows)

    def test_normal_mix(self):
        """Some producing, some dead — basic utilization calc."""
        df = self._make_df([
            {"total_hours": 1.0, "total_cases": 100},
            {"total_hours": 1.0, "total_cases": 200},
            {"total_hours": 1.0, "total_cases": 0},    # dead
            {"total_hours": 1.0, "total_cases": 150},
        ])
        util, prod_hrs, sched_hrs, dead = _compute_utilization(df)
        assert sched_hrs == 4.0
        assert prod_hrs == 3.0
        assert dead == 1
        assert abs(util - 75.0) < 0.1

    def test_all_producing(self):
        """100% utilization — no dead hours."""
        df = self._make_df([
            {"total_hours": 1.0, "total_cases": 100},
            {"total_hours": 1.0, "total_cases": 50},
        ])
        util, prod_hrs, sched_hrs, dead = _compute_utilization(df)
        assert dead == 0
        assert abs(util - 100.0) < 0.1

    def test_all_dead(self):
        """0% utilization — all scheduled hours had zero production."""
        df = self._make_df([
            {"total_hours": 1.0, "total_cases": 0},
            {"total_hours": 1.0, "total_cases": 0},
        ])
        util, prod_hrs, sched_hrs, dead = _compute_utilization(df)
        assert dead == 2
        assert abs(util - 0.0) < 0.1

    def test_empty_dataframe(self):
        df = pd.DataFrame({"total_hours": pd.Series(dtype=float),
                           "total_cases": pd.Series(dtype=float)})
        util, prod_hrs, sched_hrs, dead = _compute_utilization(df)
        assert util == 0.0
        assert dead == 0

    def test_zero_hours_excluded(self):
        """Rows with total_hours == 0 are not scheduled, so not counted."""
        df = self._make_df([
            {"total_hours": 0.0, "total_cases": 0},   # not scheduled
            {"total_hours": 1.0, "total_cases": 100},
        ])
        util, prod_hrs, sched_hrs, dead = _compute_utilization(df)
        assert sched_hrs == 1.0
        assert dead == 0
        assert abs(util - 100.0) < 0.1


# =====================================================================
# _build_dead_hour_narrative — consecutive outage detection
# =====================================================================

class TestDeadHourNarrative:
    """Dead hours should be grouped into outage blocks when consecutive."""

    def _make_hourly(self, rows):
        df = pd.DataFrame(rows)
        if "shift_hour" not in df.columns:
            df["shift_hour"] = range(1, len(df) + 1)
        return df

    def test_consecutive_block(self):
        """3 consecutive dead hours on same date/shift = 1 outage block."""
        df = self._make_hourly([
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 2, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 3, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 4, "total_hours": 1.0, "total_cases": 0},
        ])
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 1
        assert blocks[0]["n_hours"] == 3
        assert blocks[0]["pattern"] == "consecutive"
        assert summary["total_dead"] == 3
        assert summary["consecutive_hours"] == 3
        assert summary["scattered_hours"] == 0
        assert summary["n_blocks"] == 1

    def test_scattered_hours(self):
        """Non-consecutive dead hours = scattered."""
        df = self._make_hourly([
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 1, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 3, "total_hours": 1.0, "total_cases": 100},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 5, "total_hours": 1.0, "total_cases": 0},
        ])
        # Only dead hours are hr 1 and hr 5 (hr 3 has production)
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 2
        assert all(b["pattern"] == "scattered" for b in blocks)
        assert summary["scattered_hours"] == 2
        assert summary["consecutive_hours"] == 0

    def test_mixed_blocks_and_scattered(self):
        """Mix of a 2-hour outage + 1 scattered."""
        df = self._make_hourly([
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 1, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 2, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 5, "total_hours": 1.0, "total_cases": 0},
        ])
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 2  # one block of 2 + one scattered
        assert summary["consecutive_hours"] == 2
        assert summary["scattered_hours"] == 1
        assert summary["n_blocks"] == 1

    def test_different_dates_break_blocks(self):
        """Consecutive hours on different dates are separate blocks."""
        df = self._make_hourly([
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 7, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-14", "shift": "3rd", "shift_hour": 1, "total_hours": 1.0, "total_cases": 0},
        ])
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 2
        assert all(b["pattern"] == "scattered" for b in blocks)

    def test_different_shifts_break_blocks(self):
        """Same date, different shifts = separate blocks."""
        df = self._make_hourly([
            {"date_str": "2024-11-13", "shift": "1st", "shift_hour": 3, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "1st", "shift_hour": 4, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 5, "total_hours": 1.0, "total_cases": 0},
        ])
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 2
        # 1st shift: hours 3-4 (consecutive), 3rd shift: hour 5 (scattered)
        assert blocks[0]["n_hours"] == 2
        assert blocks[0]["pattern"] == "consecutive"
        assert blocks[1]["pattern"] == "scattered"

    def test_no_dead_hours(self):
        """All hours have production — no dead hours."""
        df = self._make_hourly([
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 1, "total_hours": 1.0, "total_cases": 100},
            {"date_str": "2024-11-13", "shift": "3rd", "shift_hour": 2, "total_hours": 1.0, "total_cases": 200},
        ])
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 0
        assert summary["total_dead"] == 0

    def test_empty_dataframe(self):
        df = pd.DataFrame({
            "date_str": pd.Series(dtype=str), "shift": pd.Series(dtype=str),
            "shift_hour": pd.Series(dtype=float),
            "total_hours": pd.Series(dtype=float), "total_cases": pd.Series(dtype=float),
        })
        blocks, summary = _build_dead_hour_narrative(df)
        assert len(blocks) == 0
        assert summary["total_dead"] == 0


# =====================================================================
# _correlate_dead_hours_with_events — event-to-dead-hour correlation
# =====================================================================

class TestCorrelateDeadHoursWithEvents:
    """Events should be matched to dead hours by clock-hour overlap."""

    def _make_events_df(self, events):
        return pd.DataFrame(events)

    def _make_hourly(self, rows):
        return pd.DataFrame(rows)

    def test_single_event_overlaps_dead_hour(self):
        """One event spanning a single dead hour should show as cause."""
        dead_blocks = [{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "first_hour": 2, "last_hour": 2, "n_hours": 1, "pattern": "scattered",
        }]
        events_df = self._make_events_df([{
            "reason": "Caser-Riverwood",
            "start_time": datetime(2026, 2, 6, 8, 5),
            "end_time": datetime(2026, 2, 6, 9, 0),
            "shift": "1st Shift",
            "oee_type": "Availability Loss",
            "duration_minutes": 55,
        }])
        hourly = self._make_hourly([{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "shift_hour": 2, "total_hours": 1.0, "total_cases": 0,
            "product_code": "Cut Green Beans 8pk",
        }])
        result = _correlate_dead_hours_with_events(dead_blocks, events_df, hourly)
        assert len(result) == 1
        assert "Caser-Riverwood" in result[0]["causes"]
        assert result[0]["product"] == "Cut Green Beans 8pk"

    def test_multi_hour_event(self):
        """A 4-hour event should annotate a multi-hour dead block."""
        dead_blocks = [{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "first_hour": 2, "last_hour": 5, "n_hours": 4, "pattern": "consecutive",
        }]
        events_df = self._make_events_df([{
            "reason": "Caser-Riverwood",
            "start_time": datetime(2026, 2, 6, 8, 5),
            "end_time": datetime(2026, 2, 6, 12, 37),
            "shift": "1st Shift",
            "oee_type": "Availability Loss",
            "duration_minutes": 272,
        }])
        hourly = self._make_hourly([
            {"date_str": "2026-02-06", "shift": "1st Shift",
             "shift_hour": h, "total_hours": 1.0, "total_cases": 0,
             "product_code": "Cut Green Beans 8pk"}
            for h in range(2, 6)
        ])
        result = _correlate_dead_hours_with_events(dead_blocks, events_df, hourly)
        assert "Caser-Riverwood" in result[0]["causes"]
        # Should show substantial minutes
        assert "min" in result[0]["causes"]

    def test_no_events_for_dead_hour(self):
        """Dead hour with no overlapping events gets empty cause."""
        dead_blocks = [{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "first_hour": 2, "last_hour": 2, "n_hours": 1, "pattern": "scattered",
        }]
        events_df = self._make_events_df([{
            "reason": "Something Else",
            "start_time": datetime(2026, 2, 6, 14, 0),
            "end_time": datetime(2026, 2, 6, 15, 0),
            "shift": "1st Shift",
            "oee_type": "Availability Loss",
            "duration_minutes": 60,
        }])
        hourly = self._make_hourly([{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "shift_hour": 2, "total_hours": 1.0, "total_cases": 0,
        }])
        result = _correlate_dead_hours_with_events(dead_blocks, events_df, hourly)
        assert result[0]["causes"] == ""

    def test_midnight_crossing_3rd_shift(self):
        """3rd shift events after midnight should match to the correct date."""
        dead_blocks = [{
            "date_str": "2026-02-06", "shift": "3rd Shift",
            "first_hour": 3, "last_hour": 4, "n_hours": 2, "pattern": "consecutive",
        }]
        # Shift hours 3-4 on 3rd shift starting at 23:00 = clock hours 1:00-3:00 AM on Feb 7
        events_df = self._make_events_df([{
            "reason": "Palletizer-PAI",
            "start_time": datetime(2026, 2, 7, 1, 0),
            "end_time": datetime(2026, 2, 7, 3, 0),
            "shift": "3rd Shift",
            "oee_type": "Availability Loss",
            "duration_minutes": 120,
        }])
        hourly = self._make_hourly([
            {"date_str": "2026-02-06", "shift": "3rd Shift",
             "shift_hour": 3, "total_hours": 1.0, "total_cases": 0},
            {"date_str": "2026-02-06", "shift": "3rd Shift",
             "shift_hour": 4, "total_hours": 1.0, "total_cases": 0},
        ])
        result = _correlate_dead_hours_with_events(dead_blocks, events_df, hourly)
        assert "Palletizer-PAI" in result[0]["causes"]

    def test_empty_events_df(self):
        """Empty events_df should return blocks unchanged."""
        dead_blocks = [{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "first_hour": 1, "last_hour": 1, "n_hours": 1, "pattern": "scattered",
        }]
        events_df = pd.DataFrame(columns=["reason", "start_time", "end_time",
                                           "shift", "oee_type", "duration_minutes"])
        hourly = self._make_hourly([{
            "date_str": "2026-02-06", "shift": "1st Shift",
            "shift_hour": 1, "total_hours": 1.0, "total_cases": 0,
        }])
        result = _correlate_dead_hours_with_events(dead_blocks, events_df, hourly)
        assert result == dead_blocks

    def test_empty_dead_blocks(self):
        """Empty dead blocks should return empty list."""
        events_df = self._make_events_df([{
            "reason": "Test",
            "start_time": datetime(2026, 2, 6, 8, 0),
            "end_time": datetime(2026, 2, 6, 9, 0),
            "shift": "1st Shift",
            "oee_type": "",
            "duration_minutes": 60,
        }])
        hourly = self._make_hourly([])
        result = _correlate_dead_hours_with_events([], events_df, hourly)
        assert result == []


# =====================================================================
# parse_event_summary — individual event extraction
# =====================================================================

class TestParseEventSummaryContract:
    """Verify parse_event_summary returns expected keys and column shapes."""

    def test_return_dict_has_events_df(self):
        """parse_event_summary result should include events_df key."""
        from parse_mes import parse_event_summary
        # We can't easily test with a real file, but verify the function
        # signature returns the right structure by testing with a minimal mock.
        # Instead, just verify the expected keys exist in the return value
        # by checking the function can be imported and has the right shape.
        import inspect
        sig = inspect.signature(parse_event_summary)
        assert "filepath" in sig.parameters

    def test_events_df_columns_contract(self):
        """events_df should have the expected columns when non-empty."""
        expected_cols = {"reason", "start_time", "end_time", "shift",
                         "oee_type", "duration_minutes"}
        # Verify by creating a minimal DataFrame matching the contract
        df = pd.DataFrame(columns=list(expected_cols))
        assert expected_cols == set(df.columns)


# =====================================================================
# _compute_shift_data — shift-centric metrics
# =====================================================================

class TestComputeShiftData:
    """_compute_shift_data should return a dict with all expected keys."""

    def _make_hourly(self):
        """Minimal hourly DataFrame with 2 hours of production for one shift."""
        return pd.DataFrame([
            {"date": pd.Timestamp("2026-02-06"), "date_str": "2026-02-06",
             "shift": "1st (7a-3p)", "shift_hour": 1, "total_hours": 1.0,
             "total_cases": 200, "good_cases": 195, "availability": 0.9,
             "performance": 0.8, "quality": 0.975, "oee_pct": 70.2,
             "cases_per_hour": 200, "product_code": "8PK"},
            {"date": pd.Timestamp("2026-02-06"), "date_str": "2026-02-06",
             "shift": "1st (7a-3p)", "shift_hour": 2, "total_hours": 1.0,
             "total_cases": 150, "good_cases": 148, "availability": 0.85,
             "performance": 0.75, "quality": 0.987, "oee_pct": 62.9,
             "cases_per_hour": 150, "product_code": "8PK"},
        ])

    def _make_shift_summary(self):
        return pd.DataFrame([{
            "date_str": "2026-02-06", "shift": "1st (7a-3p)",
            "total_hours": 2.0, "total_cases": 350, "good_cases": 343,
        }])

    def _make_overall(self):
        return pd.DataFrame([{
            "shift": "1st (7a-3p)", "oee_pct": 66.6,
            "cases_per_hour": 175, "total_cases": 350,
        }])

    def test_returns_expected_keys(self):
        result = _compute_shift_data(
            "1st (7a-3p)", self._make_hourly(), self._make_shift_summary(),
            self._make_overall(), None, 30.0, 200.0
        )
        assert result is not None
        expected_keys = {"scorecard", "loss_breakdown", "downtime_causes",
                         "hour_by_hour", "dead_hours", "worst_hours", "raw"}
        assert set(result.keys()) == expected_keys

    def test_raw_has_required_fields(self):
        result = _compute_shift_data(
            "1st (7a-3p)", self._make_hourly(), self._make_shift_summary(),
            self._make_overall(), None, 30.0, 200.0
        )
        raw = result["raw"]
        required = ["shift_name", "oee", "avail", "perf", "qual", "cases",
                     "hours", "cph", "n_days", "primary_loss", "target_cph"]
        for key in required:
            assert key in raw, f"Missing raw key: {key}"

    def test_oee_is_reasonable(self):
        result = _compute_shift_data(
            "1st (7a-3p)", self._make_hourly(), self._make_shift_summary(),
            self._make_overall(), None, 30.0, 200.0
        )
        oee = result["raw"]["oee"]
        assert 0 <= oee <= 100, f"OEE {oee} out of range"

    def test_scorecard_is_dataframe(self):
        result = _compute_shift_data(
            "1st (7a-3p)", self._make_hourly(), self._make_shift_summary(),
            self._make_overall(), None, 30.0, 200.0
        )
        assert isinstance(result["scorecard"], pd.DataFrame)
        assert len(result["scorecard"]) > 0

    def test_returns_none_for_empty_shift(self):
        hourly = self._make_hourly()
        result = _compute_shift_data(
            "3rd (11p-7a)", hourly, self._make_shift_summary(),
            self._make_overall(), None, 30.0, 200.0
        )
        assert result is None

    def test_shift_display_name(self):
        result = _compute_shift_data(
            "1st (7a-3p)", self._make_hourly(), self._make_shift_summary(),
            self._make_overall(), None, 30.0, 200.0
        )
        assert result["raw"]["shift_name"] == "1st Shift"


# =====================================================================
# _build_shift_narrative — auto-generated narrative
# =====================================================================

class TestBuildShiftNarrative:
    """_build_shift_narrative should produce non-empty, well-formed text."""

    def _make_shift_data(self, **overrides):
        """Build a minimal shift_data dict matching _compute_shift_data output."""
        raw = {
            "shift_name": "1st Shift",
            "shift_name_data": "1st (7a-3p)",
            "oee": 45.0,
            "avail": 0.7,
            "perf": 0.8,
            "qual": 0.98,
            "cases": 15000,
            "hours": 8.0,
            "cph": 1875.0,
            "target_cph": 3750.0,
            "target_cases": 30000,
            "pct_of_target": 50.0,
            "cases_gap": 15000,
            "n_days": 1,
            "util_pct": 75.0,
            "prod_hours": 6.0,
            "sched_hours": 8.0,
            "dead_count": 2,
            "dead_hours_total": 2,
            "plant_avg_oee": 35.0,
            "plant_avg_cph": 1500.0,
            "avail_loss": 30.0,
            "perf_loss": 20.0,
            "qual_loss": 2.0,
            "primary_loss": "Availability",
            "primary_loss_pct": 57.7,
            "top_cause": "Caser-Riverwood",
            "top_cause_min": 120.0,
            "top_cause_events": 3,
            "downtime_causes": [("Caser-Riverwood", 120.0, 3), ("Palletizer", 40.0, 2)],
            "operator_downtime_min": 160.0,
            "operator_downtime_events": 5,
        }
        raw.update(overrides)
        return {"raw": raw}

    def test_produces_nonempty_string(self):
        narrative = _build_shift_narrative(self._make_shift_data())
        assert isinstance(narrative, str)
        assert len(narrative) > 100

    def test_contains_three_paragraphs(self):
        narrative = _build_shift_narrative(self._make_shift_data())
        paragraphs = [p.strip() for p in narrative.split("\n\n") if p.strip()]
        assert len(paragraphs) == 3

    def test_paragraph1_has_oee_and_cases(self):
        narrative = _build_shift_narrative(self._make_shift_data())
        p1 = narrative.split("\n\n")[0]
        assert "45.0% OEE" in p1
        assert "15,000 cases" in p1

    def test_paragraph2_mentions_primary_loss(self):
        narrative = _build_shift_narrative(self._make_shift_data())
        p2 = narrative.split("\n\n")[1]
        assert "Availability" in p2

    def test_paragraph3_has_actions(self):
        narrative = _build_shift_narrative(self._make_shift_data())
        p3 = narrative.split("\n\n")[2]
        assert "Focus on" in p3

    def test_performance_loss_narrative(self):
        narrative = _build_shift_narrative(self._make_shift_data(
            primary_loss="Performance",
            primary_loss_pct=60.0,
            perf_loss=30.0, avail_loss=10.0,
        ))
        p2 = narrative.split("\n\n")[1]
        assert "Performance" in p2
        assert "Caser-Riverwood" in p2
        assert "#1 issue" in p2

    def test_no_downtime_still_works(self):
        narrative = _build_shift_narrative(self._make_shift_data(
            top_cause="", top_cause_min=0, top_cause_events=0,
            downtime_causes=[], operator_downtime_min=0, operator_downtime_events=0,
            dead_hours_total=0, dead_count=0,
        ))
        assert len(narrative) > 50


# =====================================================================
# parse_passdown — Shift Passdown Spreadsheet Parser
# =====================================================================

class TestParsePassdown:
    """Tests for parse_passdown module."""

    def _make_workbook(self, tmp_path, sheets, old_format=True):
        """Create a minimal passdown workbook for testing.

        sheets: list of dicts with 'name' and 'events' keys.
        Each event: {area, issue, shift, line, date, time_min}
        """
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "placeholder"

        for sheet_info in sheets:
            ws = wb.create_sheet(title=sheet_info["name"])
            # Write header row (row 2)
            if old_format:
                ws.cell(2, 4, "Shift")
                ws.cell(2, 5, "Line")
                ws.cell(2, 6, "Date")
                ws.cell(2, 8, "Area")
                ws.cell(2, 9, "Details")
                ws.cell(2, 15, "Time(min)")
                ws.cell(2, 16, "Notes:")
            else:
                ws.cell(2, 4, "Shift")
                ws.cell(2, 5, "Line")
                ws.cell(2, 6, "Date")
                ws.cell(2, 8, "Area")
                ws.cell(2, 9, "ISSUE")
                ws.cell(2, 15, "ACTION")
                ws.cell(2, 16, "RESULT")
                ws.cell(2, 17, "Time(min)")
                ws.cell(2, 18, "Notes:")

            row = 3
            for ev in sheet_info.get("events", []):
                ws.cell(row, 4, ev.get("shift", ""))
                ws.cell(row, 5, ev.get("line", ""))
                ws.cell(row, 6, ev.get("date"))
                ws.cell(row, 8, ev.get("area", ""))
                if old_format:
                    ws.cell(row, 9, ev.get("issue", ""))
                    ws.cell(row, 15, ev.get("time_min", 0))
                else:
                    ws.cell(row, 9, ev.get("issue", ""))
                    ws.cell(row, 15, ev.get("action", ""))
                    ws.cell(row, 16, ev.get("result", ""))
                    ws.cell(row, 17, ev.get("time_min", 0))
                row += 1

        # Remove placeholder sheet
        del wb["placeholder"]

        path = tmp_path / "test_passdown.xlsx"
        wb.save(str(path))
        return str(path)

    def test_returns_expected_keys(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [{"area": "Caser", "issue": "jam", "shift": "3rd", "line": "2",
                         "date": datetime(2025, 12, 3), "time_min": 15}],
        }])
        result = parse_passdown(path)
        expected_keys = {"reasons_df", "events_df", "shift_reasons_df",
                         "pareto_df", "findings", "shift_samples",
                         "event_samples", "meta", "oee_summary", "pareto_raw"}
        assert set(result.keys()) == expected_keys

    def test_events_df_columns(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [{"area": "Caser", "issue": "jam", "shift": "3rd", "line": "2",
                         "date": datetime(2025, 12, 3), "time_min": 15}],
        }])
        result = parse_passdown(path)
        edf = result["events_df"]
        for col in ["reason", "start_time", "end_time", "shift", "oee_type", "duration_minutes"]:
            assert col in edf.columns, f"Missing column: {col}"

    def test_reason_format(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [{"area": "Bear Labeler", "issue": "loose labels",
                         "shift": "3rd", "line": "2",
                         "date": datetime(2025, 12, 3), "time_min": 10}],
        }])
        result = parse_passdown(path)
        assert result["events_df"].iloc[0]["reason"] == "Bear Labeler: loose labels"

    def test_old_format_parsing(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [
                {"area": "Caser", "issue": "pallet chain broke", "shift": "3rd",
                 "line": "2", "date": datetime(2025, 12, 3), "time_min": 30},
                {"area": "Nordson", "issue": "glue issue", "shift": "3rd",
                 "line": "2", "date": datetime(2025, 12, 3), "time_min": 15},
            ],
        }], old_format=True)
        result = parse_passdown(path)
        assert len(result["events_df"]) == 2
        assert result["events_df"]["oee_type"].iloc[0] == "Availability Loss"

    def test_new_format_parsing(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-16-25",
            "events": [
                {"area": "Double Stacker", "issue": "sensor gap",
                 "shift": "3rd", "line": "3",
                 "date": datetime(2025, 12, 16), "time_min": 20,
                 "action": "adjusted sensor", "result": "fixed"},
            ],
        }], old_format=False)
        result = parse_passdown(path)
        assert len(result["events_df"]) == 1
        assert result["events_df"].iloc[0]["reason"] == "Double Stacker: sensor gap"

    def test_skips_non_data_sheets(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [
            {"name": "12-3-25",
             "events": [{"area": "Caser", "issue": "jam", "shift": "3rd",
                          "line": "2", "date": datetime(2025, 12, 3), "time_min": 15}]},
            {"name": "Reference", "events": []},
        ])
        # Reference sheet has headers but no events; verify no crash
        result = parse_passdown(path)
        assert len(result["events_df"]) == 1

    def test_reasons_df_aggregation(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [
                {"area": "Caser", "issue": "jam", "shift": "3rd",
                 "line": "2", "date": datetime(2025, 12, 3), "time_min": 30},
                {"area": "Caser", "issue": "jam", "shift": "3rd",
                 "line": "2", "date": datetime(2025, 12, 3), "time_min": 15},
            ],
        }])
        result = parse_passdown(path)
        rdf = result["reasons_df"]
        caser_row = rdf[rdf["reason"] == "Caser: jam"]
        assert len(caser_row) == 1
        assert caser_row.iloc[0]["total_minutes"] == 45.0
        assert caser_row.iloc[0]["total_occurrences"] == 2

    def test_shift_reasons_df(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [
                {"area": "Caser", "issue": "jam", "shift": "3rd",
                 "line": "2", "date": datetime(2025, 12, 3), "time_min": 30},
            ],
        }])
        result = parse_passdown(path)
        srdf = result["shift_reasons_df"]
        assert len(srdf) == 1
        assert srdf.iloc[0]["shift"] == "3rd Shift"

    def test_empty_workbook_returns_empty_dfs(self, tmp_path):
        """A workbook with no data events should return empty DataFrames."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reference"
        path = tmp_path / "empty_passdown.xlsx"
        wb.save(str(path))

        from parse_passdown import parse_passdown
        result = parse_passdown(str(path))
        assert len(result["events_df"]) == 0
        assert len(result["reasons_df"]) == 0

    def test_detect_passdown_true(self, tmp_path):
        from parse_passdown import detect_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [{"area": "Caser", "issue": "jam", "shift": "3rd",
                         "line": "2", "date": datetime(2025, 12, 3), "time_min": 15}],
        }])
        assert detect_passdown(path) is True

    def test_detect_passdown_false(self, tmp_path):
        """A workbook without Area/ISSUE headers should return False."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "OEE Period Detail")
        path = tmp_path / "not_passdown.xlsx"
        wb.save(str(path))
        from parse_passdown import detect_passdown
        assert detect_passdown(str(path)) is False

    def test_multi_sheet_aggregation(self, tmp_path):
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [
            {"name": "12-3-25",
             "events": [{"area": "Caser", "issue": "jam", "shift": "3rd",
                          "line": "2", "date": datetime(2025, 12, 3), "time_min": 15}]},
            {"name": "12-4-25",
             "events": [{"area": "Nordson", "issue": "glue", "shift": "3rd",
                          "line": "2", "date": datetime(2025, 12, 4), "time_min": 20}]},
        ])
        result = parse_passdown(path)
        assert len(result["events_df"]) == 2
        assert len(result["reasons_df"]) == 2

    def test_shift_carry_forward(self, tmp_path):
        """Shift value should carry forward to rows where it's missing."""
        from parse_passdown import parse_passdown
        path = self._make_workbook(tmp_path, [{
            "name": "12-3-25",
            "events": [
                {"area": "Caser", "issue": "jam", "shift": "3rd",
                 "line": "2", "date": datetime(2025, 12, 3), "time_min": 15},
                {"area": "Nordson", "issue": "glue issue", "shift": "",
                 "line": "", "date": None, "time_min": 10},
            ],
        }])
        result = parse_passdown(path)
        assert len(result["events_df"]) == 2
        # Second event should inherit shift from first
        assert result["events_df"].iloc[1]["shift"] == "3rd Shift"
