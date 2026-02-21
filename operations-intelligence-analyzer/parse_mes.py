"""
Parser for raw MES OEE exports.
Converts 'OEE Period Detail' and 'Event Summary (Date)' exports
into the DataFrame format expected by analyze.py.

Sheet2 block layout (13 rows each):
  Row+0:  Col B=timestamp, Col D=good cases, Col E=bad, Col F=total,
          Col G=availability, Col J=performance, Col K=quality, Col N=OEE
  Row+1:  Col C="Date",       Col E=date value
  Row+2:  Col C="Duration",   Col E=HH:MM:SS
  Row+3:  Col C="Product Code", Col E=code
  Row+4:  Col C="Product Name", Col E=name
  Row+5:  Col C="Shift",      Col E=shift string
  Row+6:  Col C="Team",       Col E=team
  Row+7:  Col C="Theoretical", Col E=rated speed
  Row+8-10: Good/Bad/Total cans
  Row+11: Job
  Row+12: Notes
"""

import re
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta

SHIFT_STARTS = {"1st Shift": 7, "2nd Shift": 15, "3rd Shift": 23}
BLOCK_SIZE = 13

_LINE_RE = re.compile(r"(Line\s*\d+)", re.IGNORECASE)


def _normalize_line(raw):
    """Extract and normalize a line identifier like 'Line 2' from a string.
    Returns None if no line identifier found.
    """
    if not raw:
        return None
    m = _LINE_RE.search(str(raw))
    if not m:
        return None
    # Normalize: "line  2" → "Line 2"
    parts = m.group(1).split()
    return f"Line {parts[-1]}"


def _safe_float(val, default=0.0):
    """Convert a cell value to float, handling None, NaN, #DIV/0!, etc."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        if val != val:  # NaN check
            return default
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if s in ("", "#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NUM!"):
            return default
        try:
            return float(s)
        except ValueError:
            return default
    return default


def _parse_timestamp(ts):
    """Parse MES timestamp to Python datetime.
    Handles double-space formatting like 'Feb  6 2026  6:55AM'.
    """
    if isinstance(ts, datetime):
        return ts
    if hasattr(ts, "to_pydatetime"):
        return ts.to_pydatetime()
    if not isinstance(ts, str) or not ts.strip():
        return None
    # Normalize multiple spaces to single space
    s = re.sub(r"\s+", " ", ts.strip())
    for fmt in [
        "%b %d %Y %I:%M%p",       # "Feb 6 2026 1:00PM"
        "%b %d %Y %I:%M %p",      # "Feb 6 2026 1:00 PM"
        "%m/%d/%Y %I:%M:%S %p",   # "2/6/2026 12:37:02 PM"
        "%m/%d/%Y %I:%M %p",      # "2/6/2026 12:37 PM"
    ]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_duration_minutes(dur_str):
    """Parse HH:MM:SS string to total minutes."""
    if not dur_str:
        return 0.0
    if isinstance(dur_str, float) and dur_str != dur_str:
        return 0.0
    parts = str(dur_str).strip().split(":")
    if len(parts) == 3:
        try:
            return int(parts[0]) * 60 + int(parts[1]) + int(parts[2]) / 60
        except ValueError:
            return 0.0
    return 0.0


def _parse_duration_hours(dur_str):
    """Parse HH:MM:SS string to total hours."""
    return _parse_duration_minutes(dur_str) / 60


def _get_shift(raw):
    """Normalize shift string. Returns None for non-production."""
    if not raw:
        return None
    s = str(raw).strip()
    if "No Shift" in s or not s:
        return None
    if "1st" in s:
        return "1st Shift"
    if "2nd" in s:
        return "2nd Shift"
    if "3rd" in s:
        return "3rd Shift"
    return None


def _shift_hour(hour_of_day, shift):
    """Calculate shift hour (1-8) from hour of day and shift name."""
    start = SHIFT_STARTS.get(shift, 7)
    return ((hour_of_day - start) % 24) + 1


def _shift_date(ts, shift):
    """Get the calendar date the shift started on."""
    if shift == "3rd Shift" and ts.hour < 7:
        return (ts - timedelta(days=1)).date()
    return ts.date()


def _hour_bucket(ts):
    """Round a timestamp down to the hour for grouping."""
    return ts.replace(minute=0, second=0, microsecond=0)


def _time_block(ts):
    """Format a time block string like '1:00pm-2:00pm' from an hour-bucketed timestamp."""
    s = ts.strftime("%I:%M%p").lstrip("0").lower()
    e = (ts + timedelta(hours=1)).strftime("%I:%M%p").lstrip("0").lower()
    return f"{s}-{e}"


# ---------------------------------------------------------------------------
# OEE Period Detail Parser
# ---------------------------------------------------------------------------
def parse_oee_period_detail(filepath):
    """
    Parse a MES 'OEE Period Detail' export.

    Sheet2 has variable-length intervals (13-row blocks).
    We aggregate sub-hourly intervals into hourly buckets to match
    the DayShiftHour format expected by analyze.py.

    Returns: (hourly_df, shift_summary_df, overall_df, hour_avg_df)
    """
    # Detect sheet name via read_only openpyxl (fast), then bulk-read with pandas
    wb_ro = openpyxl.load_workbook(filepath, read_only=True)
    hourly_sheet = "Sheet2" if "Sheet2" in wb_ro.sheetnames else wb_ro.sheetnames[1] if len(wb_ro.sheetnames) > 1 else None
    wb_ro.close()
    if not hourly_sheet:
        raise ValueError("Cannot find hourly interval sheet")

    # Bulk-read all cells at once — much faster than data_only cell-by-cell
    sheet_df = pd.read_excel(filepath, sheet_name=hourly_sheet, header=None)
    data = sheet_df.values.tolist()  # list of lists, 0-indexed
    max_row = len(data)
    raw_intervals = []

    def _cell(r, c):
        """0-indexed row, 1-indexed col (matches original ws.cell convention)."""
        if r < 0 or r >= max_row:
            return None
        if c < 1 or c > len(data[r]):
            return None
        return data[r][c - 1]

    # Scan for blocks: Row+0 col B has a timestamp, data fields are in specific positions
    row = 0  # 0-indexed
    while row + BLOCK_SIZE - 1 < max_row:
        ts_raw = _cell(row, 2)  # Col B: timestamp or label
        if ts_raw is None or (isinstance(ts_raw, float) and ts_raw != ts_raw) or str(ts_raw).strip() == "":
            row += 1
            continue

        # Skip header rows (e.g. "OEE", "Intervals")
        ts = _parse_timestamp(ts_raw)
        if ts is None:
            row += 1
            continue

        # Read interval data from block
        good_cases = _safe_float(_cell(row, 4))       # Col D: good cases
        avail      = _safe_float(_cell(row, 7))       # Col G: availability
        perf       = _safe_float(_cell(row, 10))      # Col J: performance
        qual       = _safe_float(_cell(row, 11))      # Col K: quality
        oee        = _safe_float(_cell(row, 14))      # Col N: OEE

        # Detail fields are in Col E (column 5)
        dur_str    = _cell(row + 2, 5)                # Row+2: duration
        shift_raw  = _cell(row + 5, 5)                # Row+5: shift
        product    = _cell(row + 4, 5)                # Row+4: product name
        job_str    = _cell(row + 11, 5)               # Row+11: job (contains line id)

        shift = _get_shift(shift_raw)
        if shift is None:
            row += BLOCK_SIZE
            continue

        dur_hours = _parse_duration_hours(dur_str)
        if dur_hours <= 0:
            row += BLOCK_SIZE
            continue

        bucket = _hour_bucket(ts)
        sd = _shift_date(ts, shift)

        raw_intervals.append({
            "timestamp": ts,
            "hour_bucket": bucket,
            "shift_date": sd,
            "shift": shift,
            "line": _normalize_line(job_str),
            "product": str(product).strip() if product else "",
            "good_cases": good_cases,
            "dur_hours": dur_hours,
            "availability": avail,
            "performance": perf,
            "quality": qual,
            "oee": oee,
        })

        row += BLOCK_SIZE

    if not raw_intervals:
        raise ValueError("No production intervals found in OEE Period Detail file")

    raw_df = pd.DataFrame(raw_intervals)
    # Fill missing line values — use the most common non-null line in the file
    if raw_df["line"].notna().any():
        file_line = raw_df["line"].dropna().mode().iloc[0]
        raw_df["line"] = raw_df["line"].fillna(file_line)
    else:
        raw_df["line"] = "All"

    # --- Aggregate into hourly buckets ---
    hourly_agg = (
        raw_df.groupby(["shift_date", "shift", "hour_bucket", "line"])
        .apply(_aggregate_hour, include_groups=False)
        .reset_index()
    )

    # Build the hourly DataFrame matching analyze.py expectations
    hourly = hourly_agg.copy()
    hourly["shift_hour"] = hourly.apply(
        lambda r: _shift_hour(r["hour_bucket"].hour, r["shift"]), axis=1
    )
    hourly["time_block"] = hourly["hour_bucket"].apply(_time_block)
    hourly["block_start"] = hourly["hour_bucket"]
    hourly["block_end"] = hourly["hour_bucket"] + timedelta(hours=1)
    hourly["intervals"] = 1
    hourly["date"] = pd.to_datetime(hourly["shift_date"])
    hourly["date_str"] = hourly["date"].dt.strftime("%Y-%m-%d")
    hourly["day_of_week"] = hourly["date"].dt.day_name()

    # Drop entire shift-days with zero production (e.g. 3rd shift doesn't run Saturday)
    shift_day_cases = hourly.groupby(["shift_date", "shift"])["total_cases"].transform("sum")
    hourly = hourly[shift_day_cases > 0].copy()

    # --- Shift Summary (production-weighted OEE, excluding non-production hours) ---
    hourly["_is_prod"] = (hourly["availability"] > 0) | (hourly["total_cases"] > 0)
    hourly["_prod_hours"] = hourly["total_hours"] * hourly["_is_prod"]
    hourly["_w_oee"] = hourly["oee_pct"] * hourly["_prod_hours"]
    ss = (
        hourly.groupby(["shift_date", "shift"])
        .agg(
            total_cases=("total_cases", "sum"),
            total_hours=("total_hours", "sum"),
            _prod_hours=("_prod_hours", "sum"),
            _w_oee=("_w_oee", "sum"),
            hour_blocks=("intervals", "sum"),
        )
        .reset_index()
    )
    ss["oee_pct"] = (ss["_w_oee"] / ss["_prod_hours"].replace(0, np.nan)).fillna(0)
    ss.drop(columns=["_w_oee", "_prod_hours"], inplace=True)
    ss["cases_per_hour"] = ss["total_cases"] / ss["total_hours"].replace(0, np.nan)
    ss["date"] = pd.to_datetime(ss["shift_date"])
    ss["date_str"] = ss["date"].dt.strftime("%Y-%m-%d")
    shift_summary = ss

    # --- Overall (production-weighted OEE) ---
    ov = (
        hourly.groupby("shift")
        .agg(
            total_cases=("total_cases", "sum"),
            total_hours=("total_hours", "sum"),
            _prod_hours=("_prod_hours", "sum"),
            _w_oee=("_w_oee", "sum"),
            n_intervals=("intervals", "sum"),
        )
        .reset_index()
    )
    ov["oee_pct"] = (ov["_w_oee"] / ov["_prod_hours"].replace(0, np.nan)).fillna(0)
    ov.drop(columns=["_w_oee", "_prod_hours"], inplace=True)
    ov["cases_per_hour"] = ov["total_cases"] / ov["total_hours"].replace(0, np.nan)
    overall = ov[["shift", "cases_per_hour", "oee_pct", "total_cases", "total_hours", "n_intervals"]].copy()

    # --- Hour Average (production-weighted OEE) ---
    hourly["_w_cph"] = hourly["cases_per_hour"] * hourly["_prod_hours"]
    ha = (
        hourly.groupby(["shift", "shift_hour"])
        .agg(
            time_block=("time_block", "first"),
            _w_cph=("_w_cph", "sum"),
            _w_oee=("_w_oee", "sum"),
            _prod_hours=("_prod_hours", "sum"),
            total_hours=("total_hours", "sum"),
        )
        .reset_index()
    )
    ha["cases_per_hour"] = (ha["_w_cph"] / ha["_prod_hours"].replace(0, np.nan)).fillna(0)
    ha["oee_pct"] = (ha["_w_oee"] / ha["_prod_hours"].replace(0, np.nan)).fillna(0)
    hour_avg = ha[["shift", "shift_hour", "time_block", "cases_per_hour", "oee_pct", "total_hours"]].copy()

    hourly.drop(columns=["_w_oee", "_w_cph", "_is_prod", "_prod_hours"], inplace=True, errors="ignore")
    return hourly, shift_summary, overall, hour_avg


def _aggregate_hour(group):
    """Aggregate sub-hourly intervals into one hourly record."""
    total_dur = group["dur_hours"].sum()
    total_cases = group["good_cases"].sum()

    if total_dur > 0:
        # Duration-weighted averages for OEE components
        w = group["dur_hours"]
        w_avail = (group["availability"] * w).sum() / total_dur
        w_perf = (group["performance"] * w).sum() / total_dur
        w_qual = (group["quality"] * w).sum() / total_dur
        w_oee = (group["oee"] * w).sum() / total_dur
        cph = total_cases / total_dur
    else:
        w_avail = w_perf = w_qual = w_oee = cph = 0.0

    # Most common product in this hour (by duration)
    product_code = ""
    if "product" in group.columns:
        prods = group[group["product"].astype(str).str.strip() != ""]
        if len(prods) > 0:
            product_code = prods.groupby("product")["dur_hours"].sum().idxmax()

    return pd.Series({
        "total_cases": total_cases,
        "total_hours": total_dur,
        "cases_per_hour": cph,
        "availability": w_avail,
        "performance": w_perf,
        "quality": w_qual,
        "oee_pct": w_oee * 100,  # decimal → percentage
        "product_code": product_code,
    })


# ---------------------------------------------------------------------------
# Event Summary Parser
# ---------------------------------------------------------------------------
def parse_event_summary(filepath):
    """
    Parse a MES 'Event Summary (Date)' export.

    Hierarchical structure:
      Row 6: Line total (Col B = "Line 2 - Flex")
      Rows 7+: Reason code groups (Col C) with aggregate stats,
               followed by individual events (Col D has start time).

    Returns: downtime dict matching the format expected by analyze.analyze().
    Includes:
      - reasons_df: aggregate totals per reason code
      - events_df: individual timestamped events with reason, shift, duration
      - shift_reasons_df: events grouped by (shift, reason) for shift-level Pareto
    """
    # Bulk-read all cells at once — much faster than openpyxl data_only cell-by-cell
    sheet_df = pd.read_excel(filepath, sheet_name=0, header=None)
    data = sheet_df.values.tolist()  # 0-indexed
    max_row = len(data)

    def _cell(r, c):
        """0-indexed row, 1-indexed col."""
        if r < 0 or r >= max_row:
            return None
        if c < 1 or c > len(data[r]):
            return None
        v = data[r][c - 1]
        if isinstance(v, float) and v != v:  # NaN
            return None
        return v

    reasons = []
    events = []
    current_reason = None
    detected_line = None

    for r in range(5, max_row):  # 0-indexed row 5 = original row 6
        col_b = _cell(r, 2)   # Line name (row 6 only)
        col_c = _cell(r, 3)   # Reason code group
        col_d = _cell(r, 4)   # Start time (individual events)
        col_e = _cell(r, 5)   # End time
        col_f = _cell(r, 6)   # Shift
        col_g = _cell(r, 7)   # OEE type (Availability Loss, etc.)
        col_j = _cell(r, 10)  # Count
        col_n = _cell(r, 14)  # Total duration

        # Capture line identifier from Line total row, then skip it
        if col_b and "Line" in str(col_b):
            if detected_line is None:
                detected_line = _normalize_line(col_b)
            continue

        # Reason group header: Col C has value, Col D is empty
        if col_c and not col_d:
            name = str(col_c).strip()
            if not name:
                continue
            current_reason = name
            dur = _parse_duration_minutes(col_n)
            cnt = int(_safe_float(col_j)) if col_j else 0
            if cnt > 0 or dur > 0:
                reasons.append({
                    "reason": name,
                    "total_minutes": round(dur, 1),
                    "total_occurrences": cnt,
                    "total_hours": round(dur / 60, 1),
                })
            continue

        # Individual event row: Col D has a start time
        if col_d and current_reason:
            start = _parse_timestamp(col_d)
            end = _parse_timestamp(col_e)
            if start is None:
                continue
            shift = _get_shift(col_f)
            oee_type = str(col_g).strip() if col_g else ""
            dur = _parse_duration_minutes(col_n)

            events.append({
                "reason": current_reason,
                "start_time": start,
                "end_time": end,
                "shift": shift or "",
                "oee_type": oee_type,
                "duration_minutes": round(dur, 1),
            })

    reasons_df = (
        pd.DataFrame(reasons) if reasons
        else pd.DataFrame(columns=["reason", "total_minutes", "total_occurrences", "total_hours"])
    )

    events_df = (
        pd.DataFrame(events) if events
        else pd.DataFrame(columns=["reason", "start_time", "end_time", "shift",
                                    "oee_type", "duration_minutes"])
    )

    # Build shift-level reason aggregates for per-shift Pareto
    shift_reasons_df = pd.DataFrame()
    if len(events_df) > 0 and events_df["shift"].str.strip().ne("").any():
        shift_reasons_df = (
            events_df[events_df["shift"].str.strip() != ""]
            .groupby(["shift", "reason"])
            .agg(
                total_minutes=("duration_minutes", "sum"),
                count=("duration_minutes", "count"),
            )
            .reset_index()
            .sort_values(["shift", "total_minutes"], ascending=[True, False])
        )

    return {
        "reasons_df": reasons_df,
        "events_df": events_df,
        "shift_reasons_df": shift_reasons_df,
        "pareto_df": pd.DataFrame(),
        "findings": [],
        "shift_samples": [],
        "event_samples": [],
        "meta": {},
        "oee_summary": {},
        "pareto_raw": {},
        "line": detected_line,
    }


# ---------------------------------------------------------------------------
# Format Detection
# ---------------------------------------------------------------------------
def _is_passdown(filepath):
    """Check if a file is a shift passdown spreadsheet."""
    from parse_passdown import detect_passdown
    return detect_passdown(filepath)


def detect_file_type(filepath):
    """
    Detect file format:
      'old_oee'            — processed workbook with DayShiftHour sheets
      'oee_period_detail'  — raw MES OEE Period Detail export
      'oee_overview'       — MES OEE Overview (hourly) export
      'event_overview'     — MES Event Overview export
      'event_summary'      — raw MES Event Summary (Date) export
      'passdown'           — shift passdown spreadsheet (operator notes)
      'unknown'
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames

        if "DayShiftHour" in names:
            wb.close()
            return "old_oee"

        if any("Event Summary" in n for n in names):
            wb.close()
            return "event_summary"

        ws = wb[names[0]]
        b1 = ws.cell(1, 2).value
        a1 = ws.cell(1, 1).value

        # Check for MES OEE Overview / Event Overview format:
        # Row 1: "Start" | <datetime>, Row 3: "System" | <line name>,
        # Row 4: "KPI Calc" | <kpi name>
        if str(a1).strip().lower() == "start":
            # Traksys-style pivot export — check if OEE or Event
            a4 = ws.cell(4, 1).value
            b4 = ws.cell(4, 2).value
            # Check headers further down for event columns
            # OEE Overview has columns like GroupValue, SeriesValue, Value
            # Event Overview has columns like EventID, StartDateTimeOffset, DurationSeconds
            header_row = None
            for row_idx in range(1, min(ws.max_row or 20, 20) + 1):
                cell_val = str(ws.cell(row_idx, 4).value or "")
                if cell_val in ("GroupValue", "GroupLabel"):
                    header_row = row_idx
                    break
                if cell_val in ("EventID",):
                    header_row = row_idx
                    wb.close()
                    return "event_overview"
            if str(a4).strip().lower() == "kpi calc" and str(b4).strip().lower() == "oee":
                wb.close()
                return "oee_overview"
            # Fallback: if we found GroupValue header, it's OEE overview
            if header_row is not None:
                wb.close()
                return "oee_overview"
            # If column D has EventID-like headers, it's an event overview
            for row_idx in range(1, min(ws.max_row or 20, 20) + 1):
                cell_val = str(ws.cell(row_idx, 4).value or "")
                if "event" in cell_val.lower() and "id" in cell_val.lower():
                    wb.close()
                    return "event_overview"
            # Generic Traksys export with Start/End/System — treat as OEE overview
            wb.close()
            return "oee_overview"

        wb.close()

        if b1 and "OEE" in str(b1):
            return "oee_period_detail"

        if _is_passdown(filepath):
            return "passdown"

        return "unknown"
    except Exception:
        return "unknown"

