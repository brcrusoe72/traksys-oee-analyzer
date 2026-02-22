"""Tests for vigil_agent parser and tool-enabled agent."""

from __future__ import annotations

import json

import openpyxl
import pandas as pd

from vigil_agent import VigilDataParser, VigilToolAgent


def _make_passdown(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "12-3-25"
    ws.cell(2, 4, "Shift")
    ws.cell(2, 5, "Line")
    ws.cell(2, 6, "Date")
    ws.cell(2, 8, "Area")
    ws.cell(2, 9, "ISSUE")
    ws.cell(2, 17, "Time(min)")
    ws.cell(3, 4, "3")
    ws.cell(3, 5, "Line 1")
    ws.cell(3, 6, "12/03/2025")
    ws.cell(3, 8, "Caser")
    ws.cell(3, 9, "jam")
    ws.cell(3, 17, 25)
    ws.cell(4, 4, "3")
    ws.cell(4, 5, "Line 1")
    ws.cell(4, 6, "12/03/2025")
    ws.cell(4, 8, "Caser")
    ws.cell(4, 9, "sensor")
    ws.cell(4, 17, 10)
    wb.save(str(path))


def test_parse_mixed_file_types(tmp_path):
    parser = VigilDataParser()

    csv_path = tmp_path / "a.csv"
    pd.DataFrame([{"x": 1}, {"x": 2}]).to_csv(csv_path, index=False)
    csv_art = parser.parse_file(csv_path)
    assert csv_art.kind == "csv"
    assert len(csv_art.frames["data"]) == 2

    json_path = tmp_path / "a.json"
    json_path.write_text(json.dumps([{"a": 1}, {"a": 2}]), encoding="utf-8")
    json_art = parser.parse_file(json_path)
    assert json_art.kind == "json"
    assert len(json_art.frames["data"]) == 2

    txt_path = tmp_path / "a.txt"
    txt_path.write_text("line1\nline2", encoding="utf-8")
    txt_art = parser.parse_file(txt_path)
    assert txt_art.kind == "txt"
    assert txt_art.meta["line_count"] == 2

    xlsx_path = tmp_path / "generic.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as writer:
        pd.DataFrame([{"y": 3}]).to_excel(writer, sheet_name="SheetA", index=False)
    xlsx_art = parser.parse_file(xlsx_path)
    assert xlsx_art.kind == "excel_generic"
    assert "SheetA" in xlsx_art.frames


def test_scan_and_summarize_dataset(tmp_path):
    agent = VigilToolAgent()
    (tmp_path / "note.txt").write_text("hello", encoding="utf-8")
    pd.DataFrame([{"x": 1}]).to_csv(tmp_path / "data.csv", index=False)

    scan = agent.tool_scan_directory(tmp_path)
    assert scan["file_count"] == 2
    assert scan["by_extension"][".csv"] == 1
    assert scan["by_extension"][".txt"] == 1

    summary = agent.tool_summarize_dataset(tmp_path)
    assert summary["parsed_files"] == 2
    assert summary["kinds"]["csv"] == 1
    assert summary["kinds"]["txt"] == 1


def test_query_top_downtime_causes(tmp_path):
    passdown_path = tmp_path / "passdown.xlsx"
    _make_passdown(passdown_path)
    agent = VigilToolAgent()

    result = agent.tool_query("top downtime causes", tmp_path)
    assert "top_downtime_causes" in result
    assert len(result["top_downtime_causes"]) >= 1
    assert result["top_downtime_causes"][0]["reason"].startswith("Caser")


def test_run_routes_to_tool(tmp_path):
    pd.DataFrame([{"x": 1}]).to_csv(tmp_path / "data.csv", index=False)
    agent = VigilToolAgent()

    routed = agent.run("summarize this dataset", tmp_path)
    assert routed["tool"] == "summarize_dataset"
    assert routed["result"]["parsed_files"] == 1


def test_parse_event_overview_shape(tmp_path):
    path = tmp_path / "event_overview.xlsx"
    frame = pd.DataFrame(
        [
            {
                "EventID": 1,
                "StartDateTimeOffset": "2026-02-20 07:00:00",
                "EndDateTimeOffset": "2026-02-20 07:10:00",
                "DurationSeconds": 600,
                "EventDefinitionName": "Downtime",
                "EventCategoryName": "Hydraulics",
                "OeeEventTypeName": "Availability Loss",
                "SystemName": "Line 1",
            },
            {
                "EventID": 2,
                "StartDateTimeOffset": "2026-02-20 07:20:00",
                "EndDateTimeOffset": "2026-02-20 07:25:00",
                "DurationSeconds": 300,
                "EventDefinitionName": "Downtime",
                "EventCategoryName": "Hydraulics",
                "OeeEventTypeName": "Availability Loss",
                "SystemName": "Line 1",
            },
        ]
    )
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        frame.to_excel(writer, sheet_name="Data", index=False)

    parser = VigilDataParser()
    art = parser.parse_file(path)
    assert art.kind == "event_overview"
    assert "reasons_df" in art.frames
    assert art.frames["reasons_df"].iloc[0]["reason"] == "Hydraulics"


def test_query_best_worst_oee_line_from_oee_overview(tmp_path):
    l1 = tmp_path / "OEE Overview_L1_hour.xlsx"
    l2 = tmp_path / "OEE Overview_L2_hour.xlsx"

    base_cols = ["GroupValue", "GroupLabel", "OeeDecimal", "IntervalSeconds"]
    pd.DataFrame(
        [{base_cols[0]: "2026-02-20 07:00:00", base_cols[1]: "x", base_cols[2]: 0.30, base_cols[3]: 3600}]
    ).to_excel(l1, index=False)
    pd.DataFrame(
        [{base_cols[0]: "2026-02-20 07:00:00", base_cols[1]: "x", base_cols[2]: 0.55, base_cols[3]: 3600}]
    ).to_excel(l2, index=False)

    agent = VigilToolAgent()
    result = agent.tool_query("best and worst oee line", tmp_path)

    assert result["answer"] == "Computed line-level OEE ranking."
    assert result["best_line"]["line"] == "Line 2"
    assert result["worst_line"]["line"] == "Line 1"
