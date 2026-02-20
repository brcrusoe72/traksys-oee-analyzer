"""
Parser for shift passdown spreadsheets.
Reads operator-entered downtime context (Area, Issue, Action, Duration)
and returns a dict matching the parse_event_summary() output format,
so it plugs directly into analyze() without changes.

Handles two format variants:
  - Old (12/3–12/15): Details column, Time(min) at col O, Notes at col P
  - New (12/16+):     ISSUE/ACTION/RESULT columns, Time(min) shifted right
Column positions are detected from the header row, not hardcoded.
"""

import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime

from parse_mes import SHIFT_STARTS, _safe_float

# Sheet names to always skip (reference/template/auxiliary)
_SKIP_SHEETS = {
    "reference", "template", "new_format_template", "template (2)",
    "damaged bar", "broken_knife_assembly", "bad cases",
    "videojet health", "db line 2", "rates",
}


def _parse_sheet_date(sheet_name):
    """Parse a date from the sheet name like '12-16-25' → datetime(2025,12,16)."""
    parts = sheet_name.strip().split("-")
    if len(parts) != 3:
        return None
    try:
        m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000
        return datetime(y, m, d)
    except (ValueError, TypeError):
        return None


def _build_col_map(ws):
    """Build a column map from the header row (row 2).

    Returns dict mapping normalized header names to 1-based column indices,
    or None if this isn't a data sheet.
    """
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(2, col).value
        if val is None:
            continue
        key = str(val).strip()
        if not key:
            continue
        col_map[key] = col

    # Must have Area + (ISSUE or Details) + Time(min) to be a data sheet
    has_area = "Area" in col_map
    has_issue = "ISSUE" in col_map or "Details" in col_map
    has_time = "Time(min)" in col_map
    if not (has_area and has_issue and has_time):
        return None
    return col_map


def _cell_str(ws, row, col):
    """Get a cell value as stripped string, or empty string."""
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _cell_date(ws, row, col):
    """Get a cell value as a datetime, or None."""
    val = ws.cell(row, col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if hasattr(val, "to_pydatetime"):
        return val.to_pydatetime()
    # Try parsing string dates
    s = str(val).strip()
    if not s:
        return None
    for fmt in ["%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _normalize_shift(raw):
    """Normalize shift value to '1st Shift' / '2nd Shift' / '3rd Shift' or None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s == "0":
        return None
    if "1" in s:
        return "1st Shift"
    if "2" in s:
        return "2nd Shift"
    if "3" in s:
        return "3rd Shift"
    return None


def _parse_sheet(ws, col_map, sheet_date):
    """Parse a single data sheet. Returns list of event dicts."""
    events = []

    # Resolve column indices
    area_col = col_map.get("Area")
    issue_col = col_map.get("ISSUE") or col_map.get("Details")
    time_col = col_map.get("Time(min)")
    shift_col = col_map.get("Shift")
    line_col = col_map.get("Line")
    date_col = col_map.get("Date")
    action_col = col_map.get("ACTION")
    result_col = col_map.get("RESULT")
    resolved_col = col_map.get("RESOLVED OR OPEN")
    notes_col = col_map.get("Notes:") or col_map.get("Notes")

    # Carry-forward state
    cur_shift = None
    cur_line = None
    cur_date = sheet_date

    for row in range(3, ws.max_row + 1):
        # Update carry-forward values from any non-empty cells
        if shift_col:
            raw_shift = _normalize_shift(ws.cell(row, shift_col).value)
            if raw_shift:
                cur_shift = raw_shift

        if line_col:
            raw_line = _cell_str(ws, row, line_col)
            if raw_line and raw_line != "0":
                cur_line = raw_line

        if date_col:
            raw_date = _cell_date(ws, row, date_col)
            if raw_date:
                cur_date = raw_date

        # Read downtime entry fields
        area = _cell_str(ws, row, area_col)
        issue = _cell_str(ws, row, issue_col)

        if not area or not issue:
            continue

        # Parse duration
        duration = _safe_float(ws.cell(row, time_col).value, 0.0)

        # Build reason string
        reason = f"{area}: {issue}"

        # Synthesize start_time from date + shift
        event_date = cur_date or sheet_date
        if event_date and cur_shift:
            start_hour = SHIFT_STARTS.get(cur_shift, 7)
            start_time = event_date.replace(hour=start_hour, minute=0, second=0, microsecond=0)
        elif event_date:
            start_time = event_date.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start_time = None

        event = {
            "reason": reason,
            "start_time": start_time,
            "end_time": np.nan,
            "shift": cur_shift or "",
            "oee_type": "Availability Loss",
            "duration_minutes": round(duration, 1),
        }

        # Stash extra context (not required by analyze but useful for enrichment)
        if action_col:
            action = _cell_str(ws, row, action_col)
            if action:
                event["action"] = action
        if result_col:
            result = _cell_str(ws, row, result_col)
            if result:
                event["result"] = result
        if resolved_col:
            resolved = _cell_str(ws, row, resolved_col)
            if resolved:
                event["resolved"] = resolved
        if notes_col:
            notes = _cell_str(ws, row, notes_col)
            if notes:
                event["notes"] = notes
        if cur_line:
            event["line"] = cur_line

        events.append(event)

    return events


def detect_passdown(filepath):
    """Quick check: is this file a shift passdown spreadsheet?

    Opens the workbook and looks for any sheet whose row 2 has both
    'Area' and ('ISSUE' or 'Details').
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            headers = set()
            for col in range(1, 30):
                val = ws.cell(2, col).value
                if val:
                    headers.add(str(val).strip())
            if "Area" in headers and ("ISSUE" in headers or "Details" in headers):
                wb.close()
                return True
        wb.close()
        return False
    except Exception:
        return False


def parse_passdown(filepath):
    """Parse a shift passdown spreadsheet.

    Returns a dict matching the parse_event_summary() output format:
      - reasons_df:       [reason, total_minutes, total_occurrences, total_hours]
      - events_df:        [reason, start_time, end_time, shift, oee_type, duration_minutes]
      - shift_reasons_df: [shift, reason, total_minutes, count]
      + compatibility stubs: pareto_df, findings, shift_samples, event_samples, meta, oee_summary, pareto_raw
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_events = []

    for name in wb.sheetnames:
        # Skip known non-data sheets
        if name.lower().strip() in _SKIP_SHEETS:
            continue

        ws = wb[name]
        col_map = _build_col_map(ws)
        if col_map is None:
            continue

        sheet_date = _parse_sheet_date(name)
        events = _parse_sheet(ws, col_map, sheet_date)
        all_events.extend(events)

    wb.close()

    # Build events_df (core columns only — extra context columns are kept)
    core_cols = ["reason", "start_time", "end_time", "shift", "oee_type", "duration_minutes"]
    if all_events:
        events_df = pd.DataFrame(all_events)
        # Ensure core columns exist
        for c in core_cols:
            if c not in events_df.columns:
                events_df[c] = np.nan
    else:
        events_df = pd.DataFrame(columns=core_cols)

    # Build reasons_df — aggregate by reason
    if len(events_df) > 0:
        reasons_agg = (
            events_df.groupby("reason")
            .agg(
                total_minutes=("duration_minutes", "sum"),
                total_occurrences=("duration_minutes", "count"),
            )
            .reset_index()
            .sort_values("total_minutes", ascending=False)
        )
        reasons_agg["total_hours"] = (reasons_agg["total_minutes"] / 60).round(1)
        reasons_agg["total_minutes"] = reasons_agg["total_minutes"].round(1)
        reasons_df = reasons_agg
    else:
        reasons_df = pd.DataFrame(columns=["reason", "total_minutes", "total_occurrences", "total_hours"])

    # Build shift_reasons_df — aggregate by (shift, reason)
    shift_reasons_df = pd.DataFrame()
    if len(events_df) > 0 and events_df["shift"].str.strip().ne("").any():
        shift_reasons_df = (
            events_df[events_df["shift"].str.strip() != ""]
            .groupby(["shift", "reason"])
            .agg(
                total_minutes=("duration_minutes", "sum"),
                count=("duration_minutes", "count"),
            )
            .reset_index()
            .sort_values(["shift", "total_minutes"], ascending=[True, False])
        )

    return {
        "reasons_df": reasons_df,
        "events_df": events_df,
        "shift_reasons_df": shift_reasons_df,
        "pareto_df": pd.DataFrame(),
        "findings": [],
        "shift_samples": [],
        "event_samples": [],
        "meta": {},
        "oee_summary": {},
        "pareto_raw": {},
    }

